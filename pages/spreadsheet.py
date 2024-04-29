import openpyxl
class SpreadSheet:
    def __init__(self,file_path):
            self.file_path=file_path
            self.wb_obj =openpyxl.load_workbook(self.file_path)
                # print(type(wb))
                # sheets=wb.sheetnames
                # print(sheets)
            self.sheet_obj =self.wb_obj.active            
            self.row = self.sheet_obj.max_row
            self.column = self.sheet_obj.max_column
        
    def print_row_by_phone_number(self, phone_number):         
                for i in range (2,self.row+1):
                    data_dict={}

                    Receiver_Phone_No = self.sheet_obj.cell(row=i, column=1).value
                    Amount = self.sheet_obj.cell(row=i, column=2).value
                    Remarks = self.sheet_obj.cell(row=i, column=3).value
                    # data_dict[key]=value
                    if (Receiver_Phone_No == phone_number):
                            # print("number found")
                        data_dict ={
                           "Receiver_Phone_No":Receiver_Phone_No,
                            "Amount":Amount,
                            "Remarks":Remarks
                            }
                        return data_dict
                        


                        


            

    
